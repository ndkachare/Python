import converters
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


from shoppingCart import find_max
# import ecommerce.shipping
from ecommerce import shipping
import random

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet =  wb['Sheet1']

    for row in range (2, sheet.max_row + 1 ):
        cell= sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values =  Reference(sheet, 
                min_row=2, 
                max_row=sheet.max_row,
                min_col= 4,
                max_col= 4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)

shipping.calc_shipping()
print(converters.kg_to_lb(100))
numbers = [10,15,50,52,20]
# max = find_max(numbers)  it a bad practice as changing the meaning of a built in function
maximum = find_max(numbers)
print(maximum)
